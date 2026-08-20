"""Parse AFAC annual workbooks and the DATATUR long-form mirror."""

from __future__ import annotations

from calendar import monthrange
from collections import defaultdict
import csv
from datetime import date
import hashlib
import json
from pathlib import Path
import re
import unicodedata

import openpyxl
import polars as pl
from pypdf import PdfReader
import xlrd

from src.common.quality import log_issue_once
from src.config import PATHS
from src.parse.sec.common import write_parquet_atomic


PARSER_VERSION = "afac_monthly_stats_v1"
ANNUAL_START_YEAR = 2015
PASSENGER_SHEETS = {"PAXREG": "scheduled", "PASREG": "scheduled", "PAXFLET": "charter", "PASFLET": "charter"}
MONTH_NAMES = {
    "ENERO": 1,
    "FEBRERO": 2,
    "MARZO": 3,
    "ABRIL": 4,
    "MAYO": 5,
    "JUNIO": 6,
    "JULIO": 7,
    "AGOSTO": 8,
    "SEPTIEMBRE": 9,
    "OCTUBRE": 10,
    "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}
PDF_INTEGER = re.compile(r"^-?\d{1,3}(?:,\d{3})*$")


def _normalize(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(character for character in text if not unicodedata.combining(character))
    return " ".join(re.sub(r"[^A-Z0-9]+", " ", text.upper()).split())


def _compact(value: object) -> str:
    return re.sub(r"[^A-Z0-9]", "", _normalize(value))


def _number(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if text in {"", "-", "N/A", "N.D."}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _rows(path: Path, sheet_name: str) -> list[list[object]]:
    if path.suffix.casefold() == ".xls":
        book = xlrd.open_workbook(path)
        sheet = book.sheet_by_name(sheet_name)
        return [sheet.row_values(index) for index in range(sheet.nrows)]
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return [list(row) for row in book[sheet_name].iter_rows(values_only=True)]
    finally:
        book.close()


def _sheet_names(path: Path) -> list[str]:
    if path.suffix.casefold() == ".xls":
        return xlrd.open_workbook(path, on_demand=True).sheet_names()
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return list(book.sheetnames)
    finally:
        book.close()


def parse_annual_passenger_workbook(
    path: Path, year: int
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    """Parse passenger facts and workbook TOTAL checks from either Excel family."""

    facts: list[dict[str, object]] = []
    checks: list[dict[str, object]] = []
    available = {name.upper(): name for name in _sheet_names(path)}
    for normalized_sheet, service_type in PASSENGER_SHEETS.items():
        actual_sheet = available.get(normalized_sheet)
        if actual_sheet is None:
            continue
        rows = _rows(path, actual_sheet)
        is_domestic_carrier: bool | None = None
        market: str | None = None
        in_data = False
        block_facts: list[dict[str, object]] = []
        block_start = 0
        for row_number, row in enumerate(rows, start=1):
            joined = _normalize(" ".join(str(value) for value in row if value not in (None, "")))
            if "EMPRESAS EXTRANJERAS" in joined or "FOREIGN AIR CARRIER" in joined:
                is_domestic_carrier = False
            elif "EMPRESAS NACIONALES" in joined or "DOMESTIC AIR CARRIER" in joined:
                is_domestic_carrier = True
            if "SERVICIO" in joined or "SERVICE" in joined:
                if "INTERNACIONAL" in joined or "INTERNATIONAL" in joined:
                    market = "international"
                elif "NACIONAL" in joined or "DOMESTIC" in joined:
                    market = "domestic"

            carrier_column = next(
                (
                    index
                    for index, value in enumerate(row)
                    if "EMPRESA" in _normalize(value) or "AIR CARRIER" in _normalize(value)
                ),
                None,
            )
            if carrier_column is not None and any(
                "JAN" in _normalize(value) or "ENE" in _normalize(value) for value in row
            ):
                if is_domestic_carrier is None or market is None:
                    raise ValueError(
                        f"Missing AFAC block context in {path.name}/{actual_sheet} row {row_number}"
                    )
                in_data = True
                block_facts = []
                block_start = row_number
                continue
            if not in_data or not row:
                continue

            carrier_name = next(
                (str(value).strip() for value in row if isinstance(value, str) and value.strip()),
                "",
            )
            if not carrier_name:
                continue
            compact_name = _compact(carrier_name)
            month_values = [_number(value) for value in row[1:13]]
            if compact_name in {"TOTAL", "TOTALTOTAL"}:
                for month, expected in enumerate(month_values, start=1):
                    if expected is None:
                        continue
                    actual = sum(
                        float(fact["value"])
                        for fact in block_facts
                        if fact["month"] == month
                    )
                    denominator = max(abs(expected), 1.0)
                    checks.append(
                        {
                            "year": year,
                            "period_id": f"{year}M{month:02d}",
                            "sheet": actual_sheet,
                            "block_start_row": block_start,
                            "is_domestic_carrier": is_domestic_carrier,
                            "service_type": service_type,
                            "market": market,
                            "actual": actual,
                            "expected": expected,
                            "difference": actual - expected,
                            "relative_difference": abs(actual - expected) / denominator,
                        }
                    )
                facts.extend(block_facts)
                block_facts = []
                in_data = False
                continue
            if _normalize(carrier_name).startswith("TOTAL "):
                continue  # Foreign-carrier regional subtotals are not airline rows.
            if not any(value is not None for value in month_values):
                continue
            for month, value in enumerate(month_values, start=1):
                if value is None:
                    continue
                block_facts.append(
                    {
                        "year": year,
                        "month": month,
                        "source_carrier_name": carrier_name,
                        "is_domestic_carrier": is_domestic_carrier,
                        "service_type": service_type,
                        "market": market,
                        "value": value,
                        "source_row_number": row_number,
                    }
                )
        if in_data and block_facts:
            raise ValueError(f"Unclosed AFAC data block in {path.name}/{actual_sheet}")
    if not facts or not checks:
        raise ValueError(f"No AFAC passenger facts or totals parsed from {path}")
    return facts, checks


def _metadata(path: Path) -> dict[str, object]:
    meta_path = path.with_suffix(path.suffix + ".meta.json")
    metadata = json.loads(meta_path.read_text(encoding="utf-8"))
    actual_hash = hashlib.sha256(path.read_bytes()).hexdigest()
    if metadata.get("sha256") != actual_hash:
        raise ValueError(f"Bronze hash mismatch for {path}")
    return metadata


def _latest_annual_workbook(year: int) -> Path:
    directory = PATHS.bronze / "afac" / str(year)
    candidates = [
        path
        for path in directory.glob("afac_afac_annual_workbook_*")
        if path.suffix.casefold() in {".xls", ".xlsx"}
    ]
    if not candidates:
        raise FileNotFoundError(f"Missing AFAC annual workbook for {year}")
    return max(candidates, key=lambda path: path.stat().st_mtime_ns)


def _load_crosswalk() -> dict[str, list[dict[str, str]]]:
    path = PATHS.data / "reference" / "carrier_crosswalk.csv"
    result: dict[str, list[dict[str, str]]] = defaultdict(list)
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            if row["source_system"] != "afac":
                continue
            result[_normalize(row["source_carrier_name"])].append(row)
    return result


def _mapped_carrier(
    crosswalk: dict[str, list[dict[str, str]]], carrier_name: str, year: int, month: int
) -> tuple[str | None, str | None]:
    period = f"{year}-{month:02d}"
    for row in crosswalk.get(_normalize(carrier_name), []):
        if row["valid_from"] and period < row["valid_from"]:
            continue
        if row["valid_to"] and period > row["valid_to"]:
            continue
        return row["carrier_key"] or None, row["iata"] or None
    return None, None


def _month_dates(year: int, month: int) -> tuple[date, date]:
    return date(year, month, 1), date(year, month, monthrange(year, month)[1])


def _annual_facts(
    crosswalk: dict[str, list[dict[str, str]]]
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    records: list[dict[str, object]] = []
    checks: list[dict[str, object]] = []
    for year in range(ANNUAL_START_YEAR, 2026):
        path = _latest_annual_workbook(year)
        metadata = _metadata(path)
        raw_facts, raw_checks = parse_annual_passenger_workbook(path, year)
        source_file = path.relative_to(PATHS.bronze).as_posix()
        family = "xlsx_wide_2012_2016" if year <= 2016 else "xlsx_modern_2017_2025"
        for raw in raw_facts:
            month = int(raw["month"])
            start, end = _month_dates(year, month)
            carrier_key, iata = _mapped_carrier(
                crosswalk, str(raw["source_carrier_name"]), year, month
            )
            records.append(
                {
                    "period_id": f"{year}M{month:02d}",
                    "period_type": "month",
                    "period_start_date": start,
                    "period_end_date": end,
                    "carrier_key": carrier_key,
                    "source_carrier_name": raw["source_carrier_name"],
                    "iata_code": iata,
                    "is_domestic_carrier": raw["is_domestic_carrier"],
                    "service_type": raw["service_type"],
                    "market": raw["market"],
                    "metric_key": "passengers",
                    "value": raw["value"],
                    "unit": "passengers",
                    "is_preliminary": year == 2025,
                    "is_estimated": False,
                    "footnote_text": None,
                    "source_system": "afac",
                    "source_file": source_file,
                    "source_hash": metadata["sha256"],
                    "ingested_at": metadata["downloaded_at"],
                    "parser_version": PARSER_VERSION,
                    "source_row_number": raw["source_row_number"],
                    "source_region": None,
                    "source_row_count": 1,
                    "source_family": family,
                }
            )
        for check in raw_checks:
            checks.append(
                {
                    **check,
                    "check_scope": "annual_workbook_total",
                    "source_file": source_file,
                    "passed": float(check["relative_difference"]) <= 0.001,
                }
            )
    return records, checks


def _estimate_rules() -> dict[tuple[str, str], str]:
    rules: dict[tuple[str, str], str] = {}
    pattern = re.compile(
        r"Datos estimados de\s+(.+?)\s+para\s+([A-Za-zÁÉÍÓÚáéíóú]+)\s+(20\d{2})",
        re.IGNORECASE,
    )
    for path in (PATHS.bronze / "afac").glob("20??/??/*.pdf"):
        text = "\n".join(page.extract_text() or "" for page in PdfReader(path).pages)
        for carrier, month_name, year in pattern.findall(text):
            month = MONTH_NAMES.get(_normalize(month_name))
            if month is None:
                continue
            footnote = f"Datos estimados de {carrier.strip()} para {month_name.strip()} {year}."
            rules[(f"{year}M{month:02d}", _normalize(carrier))] = footnote
    return rules


def _pdf_line_values(line: str) -> tuple[str, list[int]] | None:
    tokens = line.split()
    first_number = next(
        (index for index, token in enumerate(tokens) if PDF_INTEGER.fullmatch(token)),
        None,
    )
    if first_number is None:
        return None
    values = [
        int(token.replace(",", ""))
        for token in tokens[first_number:]
        if PDF_INTEGER.fullmatch(token)
    ]
    if len(values) < 6:
        return None
    return " ".join(tokens[:first_number]), values


def _pdf_2026_facts(
    crosswalk: dict[str, list[dict[str, str]]]
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    """Parse scheduled passengers from each contemporaneous 2026 DATATUR PDF."""

    records: list[dict[str, object]] = []
    checks: list[dict[str, object]] = []
    estimates = _estimate_rules()
    for month in range(1, 7):
        directory = PATHS.bronze / "afac" / "2026" / f"{month:02d}"
        path = next(directory.glob("*.pdf"), None)
        if path is None:
            raise FileNotFoundError(f"Missing DATATUR PDF for 2026M{month:02d}")
        metadata = _metadata(path)
        pages = [page.extract_text() or "" for page in PdfReader(path).pages]
        if len(pages) < 2:
            raise ValueError(f"DATATUR bulletin has fewer than two pages: {path}")
        source_file = path.relative_to(PATHS.bronze).as_posix()
        period_id = f"2026M{month:02d}"
        start, end = _month_dates(2026, month)
        for page_number, (page_text, market) in enumerate(
            ((pages[0], "domestic"), (pages[1], "international")), start=1
        ):
            page_values: list[float] = []
            expected_total: float | None = None
            region = "Mexicana" if market == "domestic" else None
            is_domestic = market == "domestic"
            for line_number, line in enumerate(page_text.splitlines(), start=1):
                parsed = _pdf_line_values(line)
                if parsed is None:
                    continue
                carrier_name, values = parsed
                compact_name = _compact(carrier_name)
                if market == "domestic" and _normalize(carrier_name).startswith("TOTAL GENERAL"):
                    expected_total = float(values[-1])
                    continue
                if market == "international" and compact_name in {"TOTAL", "TOTALTOTAL"}:
                    expected_total = float(values[-1])
                    continue
                if market == "international" and _normalize(carrier_name).startswith("TOTAL "):
                    region_label = _normalize(carrier_name).removeprefix("TOTAL ")
                    region = region_label.title()
                    is_domestic = region_label == "MEXICANAS"
                    continue
                if carrier_name in {"", "Aerolíneas", "Aerolineas"}:
                    continue
                value = float(values[-1])
                page_values.append(value)
                carrier_key, iata = _mapped_carrier(
                    crosswalk, carrier_name, 2026, month
                )
                estimated_note = next(
                    (
                        note
                        for (rule_period, carrier_fragment), note in estimates.items()
                        if rule_period == period_id
                        and carrier_fragment in _normalize(carrier_name)
                    ),
                    None,
                )
                records.append(
                    {
                        "period_id": period_id,
                        "period_type": "month",
                        "period_start_date": start,
                        "period_end_date": end,
                        "carrier_key": carrier_key,
                        "source_carrier_name": carrier_name,
                        "iata_code": iata,
                        "is_domestic_carrier": is_domestic,
                        "service_type": "scheduled",
                        "market": market,
                        "metric_key": "passengers",
                        "value": value,
                        "unit": "passengers",
                        "is_preliminary": True,
                        "is_estimated": estimated_note is not None,
                        "footnote_text": estimated_note,
                        "source_system": "afac_datatur",
                        "source_file": source_file,
                        "source_hash": metadata["sha256"],
                        "ingested_at": metadata["downloaded_at"],
                        "parser_version": PARSER_VERSION,
                        "source_row_number": line_number,
                        "source_region": region,
                        "source_row_count": 1,
                        "source_family": "datatur_monthly_bulletin_pdf",
                    }
                )
            if expected_total is None:
                raise ValueError(f"Missing PDF TOTAL in {path.name} page {page_number}")
            actual = sum(page_values)
            relative_difference = abs(actual - expected_total) / max(expected_total, 1.0)
            checks.append(
                {
                    "year": 2026,
                    "period_id": period_id,
                    "sheet": f"pdf_page_{page_number}",
                    "block_start_row": 1,
                    "is_domestic_carrier": market == "domestic",
                    "service_type": "scheduled",
                    "market": market,
                    "actual": actual,
                    "expected": expected_total,
                    "difference": actual - expected_total,
                    "relative_difference": relative_difference,
                    "check_scope": "datatur_pdf_total",
                    "source_file": source_file,
                    "passed": relative_difference <= 0.001,
                }
            )
    return records, checks


def _datatur_2026_facts(
    crosswalk: dict[str, list[dict[str, str]]]
) -> tuple[list[dict[str, object]], int]:
    candidates = list((PATHS.bronze / "afac" / "database" / "extracted").glob("*.xlsx"))
    if not candidates:
        raise FileNotFoundError("Missing extracted DATATUR database workbook")
    path = max(candidates, key=lambda item: item.stat().st_mtime_ns)
    metadata = _metadata(path)
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        rows = book["AFAC"].iter_rows(values_only=True)
        headers = [str(value) for value in next(rows)]
        frame = pl.DataFrame(
            [dict(zip(headers, row, strict=True)) for row in rows], infer_schema_length=None
        )
    finally:
        book.close()
    renamed = frame.rename(
        {
            "Año": "year",
            "Tipo": "market_raw",
            "Servicio": "service_raw",
            "Región": "region_raw",
            "Aerolinea": "source_carrier_name",
            "Id_mes": "month",
            "Pasajeros": "value",
        }
    ).with_columns(
        pl.col("market_raw", "service_raw", "region_raw", "source_carrier_name")
        .cast(pl.String)
        .str.strip_chars(),
        pl.col("year", "month").cast(pl.Int64),
        pl.col("value").cast(pl.Float64),
    )
    frame_2026 = renamed.filter(
        (pl.col("year") == 2026)
        & (_normalize_expression(pl.col("service_raw")) != "REGULAR")
    )
    key = ["year", "month", "market_raw", "service_raw", "region_raw", "source_carrier_name"]
    grouped = frame_2026.group_by(key, maintain_order=True).agg(
        pl.col("value").sum(), pl.len().alias("source_row_count")
    )
    duplicate_rows = grouped.filter(pl.col("source_row_count") > 1).height
    estimates = _estimate_rules()
    source_file = path.relative_to(PATHS.bronze).as_posix()
    records: list[dict[str, object]] = []
    for source_row_number, row in enumerate(grouped.iter_rows(named=True), start=2):
        year = int(row["year"])
        month = int(row["month"])
        period_id = f"{year}M{month:02d}"
        start, end = _month_dates(year, month)
        carrier_name = str(row["source_carrier_name"])
        carrier_key, iata = _mapped_carrier(crosswalk, carrier_name, year, month)
        estimated_note = next(
            (
                note
                for (rule_period, carrier_fragment), note in estimates.items()
                if rule_period == period_id and carrier_fragment in _normalize(carrier_name)
            ),
            None,
        )
        records.append(
            {
                "period_id": period_id,
                "period_type": "month",
                "period_start_date": start,
                "period_end_date": end,
                "carrier_key": carrier_key,
                "source_carrier_name": carrier_name,
                "iata_code": iata,
                "is_domestic_carrier": _normalize(row["region_raw"]) == "MEXICANA",
                "service_type": "scheduled" if _normalize(row["service_raw"]) == "REGULAR" else "charter",
                "market": "domestic" if _normalize(row["market_raw"]) == "NACIONAL" else "international",
                "metric_key": "passengers",
                "value": row["value"],
                "unit": "passengers",
                "is_preliminary": True,
                "is_estimated": estimated_note is not None,
                "footnote_text": estimated_note,
                "source_system": "afac_datatur",
                "source_file": source_file,
                "source_hash": metadata["sha256"],
                "ingested_at": metadata["downloaded_at"],
                "parser_version": PARSER_VERSION,
                "source_row_number": source_row_number,
                "source_region": row["region_raw"],
                "source_row_count": row["source_row_count"],
                "source_family": "datatur_long_database",
            }
        )
    return records, duplicate_rows


def _normalize_expression(expression: pl.Expr) -> pl.Expr:
    """Polars equivalent of the subset of normalization used for source labels."""

    return expression.cast(pl.String).str.strip_chars().str.to_uppercase()


def _frame(records: list[dict[str, object]]) -> pl.DataFrame:
    return pl.DataFrame(records, infer_schema_length=None).with_columns(
        pl.col("carrier_key", "iata_code", "footnote_text", "source_region").cast(pl.String),
        pl.col("value").cast(pl.Float64),
        pl.col("source_row_number", "source_row_count").cast(pl.Int64),
    ).sort(
        ["period_id", "metric_key", "market", "service_type", "source_carrier_name", "source_row_number"]
    )


def _sec_reconciliation(facts: pl.DataFrame) -> tuple[pl.DataFrame, dict[str, float | int]]:
    afac = (
        facts.filter(
            pl.col("carrier_key").is_in(["AEROMEXICO", "AEROMEXICO_CONNECT"])
            & (pl.col("metric_key") == "passengers")
        )
        .group_by("period_id")
        .agg(pl.col("value").sum().alias("afac_passengers"))
    )
    sec = (
        pl.read_parquet(PATHS.silver / "sec_operating_metrics.parquet")
        .filter(
            (pl.col("period_type") == "month")
            & (pl.col("metric_key") == "passengers")
            & (pl.col("segment") == "total")
        )
        .select("period_id", pl.col("value_normalized").alias("sec_passengers"))
    )
    joined = afac.join(sec, on="period_id", how="inner").with_columns(
        (pl.col("afac_passengers") - pl.col("sec_passengers")).alias("difference"),
        ((pl.col("afac_passengers") - pl.col("sec_passengers")) / pl.col("sec_passengers")).alias("difference_pct"),
    ).sort("period_id")
    if joined.height < 3:
        raise ValueError("Insufficient AFAC/SEC overlapping passenger months")
    correlation = float(joined.select(pl.corr("afac_passengers", "sec_passengers")).item())
    stats: dict[str, float | int] = {
        "overlap_months": joined.height,
        "correlation": correlation,
        "mean_difference": float(joined["difference"].mean()),
        "mean_difference_pct": float(joined["difference_pct"].mean()),
        "median_difference_pct": float(joined["difference_pct"].median()),
    }
    return joined.with_columns(pl.lit(PARSER_VERSION).alias("parser_version")), stats


def _write_unmapped(facts: pl.DataFrame) -> tuple[Path, int]:
    target = PATHS.quality / "afac_unmapped_carriers.csv"
    unmapped = (
        facts.filter(pl.col("carrier_key").is_null())
        .group_by("source_carrier_name")
        .agg(
            pl.col("period_id").min().alias("first_period"),
            pl.col("period_id").max().alias("last_period"),
            pl.len().alias("fact_rows"),
        )
        .sort("source_carrier_name")
    )
    target.parent.mkdir(parents=True, exist_ok=True)
    unmapped.write_csv(target)
    return target, unmapped.height


def _format_family(year: int) -> str:
    if year <= 2008:
        return "legacy_biff_wide"
    if year == 2009:
        return "biff_wide_plus_hours"
    if year <= 2011:
        return "biff_wide_plus_operational"
    if year <= 2016:
        return "ooxml_wide_plus_operational"
    if year <= 2019:
        return "ooxml_summary_plus_operational"
    if year <= 2024:
        return "ooxml_with_revision_history"
    return "ooxml_modern"


def write_inventory() -> Path:
    """Write the reproducible AFAC source and format-family inventory."""

    lines = [
        "# Inventario de fuentes AFAC",
        "",
        "Inventario generado a partir de la capa bronze. La serie anual oficial de gob.mx",
        "está completa para 1992–2025; DATATUR aporta boletines mensuales 2024M01–2026M06",
        "y una base larga 2016M01–2026M06. Todos los archivos tienen SHA-256 y metadatos",
        "inmutables en `data/bronze/_manifest.jsonl`.",
        "",
        "## Familias de formato",
        "",
        "| Familia | Periodos | Rasgos | Uso en silver |",
        "|---|---|---|---|",
        "| `legacy_biff_wide` | 1992–2008 | `.xls`, hojas PAX/PAS por servicio, bloques apilados y meses en columnas | Fixture y archivo histórico; no entra al corte analítico 2015+ |",
        "| `biff_wide_plus_hours` | 2009 | Agrega hojas de horas | Inventariada |",
        "| `biff_wide_plus_operational` | 2010–2011 | Agrega OPREG/OPFLET en formato largo | Inventariada |",
        "| `ooxml_wide_plus_operational` | 2012–2016 | Migración a `.xlsx`; PAX ancho y operación larga | Pasajeros desde 2015 |",
        "| `ooxml_summary_plus_operational` | 2017–2019 | Incorpora hoja Resumen | Pasajeros |",
        "| `ooxml_with_revision_history` | 2020–2024 | Incorpora historial explícito de revisiones | Pasajeros |",
        "| `ooxml_modern` | 2025 | Misma topología moderna; sin hoja de revisiones en la descarga actual | Pasajeros preliminares |",
        "| `datatur_monthly_bulletin_pdf` | 2024M01–2026M06 | Tablas mensuales, participación y notas de estimación | Canónico para vuelos regulares 2026 |",
        "| `datatur_long_database` | 2016M01–2026M06 | Una fila por mes/mercado/servicio/región/aerolínea | Canónico para fletamento 2026; solape usado para QA |",
        "",
        "## Archivos",
        "",
        "| Periodo | Archivo bronze | Formato | Tamaño (bytes) | Hojas/páginas | Descargado | Método | Familia |",
        "|---|---|---|---:|---|---|---|---|",
    ]
    for year in range(1992, 2026):
        path = _latest_annual_workbook(year)
        metadata = _metadata(path)
        sheets = ", ".join(_sheet_names(path))
        lines.append(
            f"| {year} | `{path.relative_to(PATHS.bronze).as_posix()}` | {path.suffix[1:]} | "
            f"{path.stat().st_size:,} | {sheets} | Sí | {metadata['download_method']} | `{_format_family(year)}` |"
        )
    for year in range(2024, 2027):
        for month in range(1, 13):
            if (year, month) > (2026, 6):
                break
            directory = PATHS.bronze / "afac" / str(year) / f"{month:02d}"
            pdf = next(directory.glob("*.pdf"), None)
            if pdf is None:
                lines.append(
                    f"| {year}M{month:02d} | — | pdf | — | — | No | — | `datatur_monthly_bulletin_pdf` |"
                )
                continue
            metadata = _metadata(pdf)
            page_count = len(PdfReader(pdf).pages)
            lines.append(
                f"| {year}M{month:02d} | `{pdf.relative_to(PATHS.bronze).as_posix()}` | pdf | "
                f"{pdf.stat().st_size:,} | {page_count} páginas | Sí | {metadata['download_method']} | `datatur_monthly_bulletin_pdf` |"
            )
    database = max(
        (PATHS.bronze / "afac" / "database" / "extracted").glob("*.xlsx"),
        key=lambda item: item.stat().st_mtime_ns,
    )
    database_metadata = _metadata(database)
    lines.append(
        f"| 2016M01–2026M06 | `{database.relative_to(PATHS.bronze).as_posix()}` | xlsx | "
        f"{database.stat().st_size:,} | AFAC | Sí | {database_metadata['download_method']} | `datatur_long_database` |"
    )
    lines.extend(
        [
            "",
            "## Precedencia de fuentes",
            "",
            "- 2015–2025: libros anuales oficiales AFAC, porque conservan bloques y filas TOTAL.",
            "- 2026 vuelos regulares: boletín mensual DATATUR/AFAC del mismo periodo.",
            "- 2026 fletamento: base larga DATATUR, ya que el boletín no publica ese desglose.",
            "- Los solapes no se concatenan: se usan para detectar revisiones y diferencias de versión.",
            "",
        ]
    )
    target = PATHS.root / "docs" / "afac-inventario.md"
    target.write_text("\n".join(lines), encoding="utf-8")
    return target


def run_afac_parse() -> dict[str, object]:
    """Build AFAC silver, strict workbook checks, and the SEC reconciliation."""

    crosswalk = _load_crosswalk()
    annual_records, total_checks = _annual_facts(crosswalk)
    pdf_records, pdf_checks = _pdf_2026_facts(crosswalk)
    datatur_records, duplicate_groups = _datatur_2026_facts(crosswalk)
    facts = _frame(annual_records + pdf_records + datatur_records)
    checks = pl.DataFrame(total_checks + pdf_checks).sort(
        ["period_id", "sheet", "block_start_row"]
    )
    failed_checks = checks.filter(~pl.col("passed"))
    if failed_checks.height:
        examples = failed_checks.head(5).select("period_id", "sheet", "relative_difference").to_dicts()
        raise ValueError(f"AFAC workbook TOTAL reconciliation failed: {examples}")

    observed_periods = facts["period_id"].unique().sort().to_list()
    expected_periods = [
        f"{year}M{month:02d}"
        for year in range(ANNUAL_START_YEAR, 2027)
        for month in range(1, 13)
        if (year, month) <= (2026, 6)
    ]
    missing_periods = sorted(set(expected_periods) - set(observed_periods))
    if missing_periods:
        raise ValueError(f"Missing AFAC monthly periods: {missing_periods}")

    write_parquet_atomic(facts, PATHS.silver / "afac_monthly_stats.parquet")
    write_parquet_atomic(checks, PATHS.quality / "afac_total_checks.parquet")
    reconciliation, reconciliation_stats = _sec_reconciliation(facts)
    write_parquet_atomic(
        reconciliation, PATHS.quality / "afac_sec_reconciliation.parquet"
    )
    if float(reconciliation_stats["correlation"]) <= 0.95:
        raise ValueError(f"AFAC/SEC passenger correlation is too low: {reconciliation_stats}")

    unmapped_path, unmapped_count = _write_unmapped(facts)
    inventory_path = write_inventory()
    for row in (
        facts.filter(pl.col("carrier_key").is_null())
        .group_by("source_carrier_name")
        .agg(pl.len().alias("rows"), pl.col("source_file").first())
        .iter_rows(named=True)
    ):
        log_issue_once(
            "silver",
            "afac_monthly_stats",
            str(row["source_file"]),
            "warning",
            "unmapped_entity",
            f"Unmapped AFAC carrier: {row['source_carrier_name']}",
            int(row["rows"]),
        )
    if duplicate_groups:
        log_issue_once(
            "bronze",
            "datatur_database_member",
            datatur_records[0]["source_file"],
            "info",
            "schema_drift",
            "DATATUR contains repeated natural keys; the parser preserves provenance in source_row_count and sums the source rows.",
            duplicate_groups,
        )

    return {
        "network_used": False,
        "silver_rows": facts.height,
        "period_min": observed_periods[0],
        "period_max": observed_periods[-1],
        "period_count": len(observed_periods),
        "missing_periods": missing_periods,
        "total_checks": checks.height,
        "failed_total_checks": failed_checks.height,
        "datatur_duplicate_groups_2026": duplicate_groups,
        "unmapped_carrier_names": unmapped_count,
        "unmapped_file": unmapped_path.relative_to(PATHS.root).as_posix(),
        "inventory_file": inventory_path.relative_to(PATHS.root).as_posix(),
        "sec_reconciliation": reconciliation_stats,
        "default_business_view": "AEROMEXICO + AEROMEXICO_CONNECT",
    }


def main() -> int:
    print(json.dumps(run_afac_parse(), indent=2, ensure_ascii=False, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
